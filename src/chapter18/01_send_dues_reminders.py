#! /usr/bin/env python3
# Sends emails based on payment status in spreadsheet.

import openpyxl
import smtplib

wb = openpyxl.load_workbook("excel/dues_records.xlsx")
sheet = wb["Sheet1"]
last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value
# Check each member's payment status.
unpaid_members = {}
for row_number in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=row_number, column=last_col).value
    if payment != "paid":
        name = sheet.cell(row=row_number, column=1).value
        email = sheet.cell(row=row_number, column=2).value
        unpaid_members[name] = email
wb.close()
print(unpaid_members)

# We use MailHog as a fake SMTP server using Docker
# https://www.linkedin.com/pulse/setting-up-smtp-mail-server-using-mailhog-docker-image-chopparapu
smtp_obj = smtplib.SMTP("localhost", 1025)
smtp_obj.ehlo()
for name, email in unpaid_members.items():
    body = f"""Subject: {latest_month} dues unpaid.

Dear {name},
Records show that you have not paid {latest_month}. Please make this payment as soon as possible. Thank you!
"""
    print(f"Sending email to {email}")
    send_email_status = smtp_obj.sendmail("my_email_address@example.com", email, body, )
    if send_email_status != {}:
        print(f"There was a problem sending a email to {email}: {send_email_status}")

smtp_obj.quit()
