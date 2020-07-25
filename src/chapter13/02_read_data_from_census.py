#! /usr/bin/env python3
# Reading Data from a Spreadsheet (2010 US Census)
import json
import pprint
import openpyxl
from openpyxl.styles import Font

print("Reading Data from a Spreadsheet (2010 US Census)")

wb = openpyxl.load_workbook("extra_files/censuspopdata.xlsx")

sheet = wb.worksheets[0]
info = {}
for row in range(2, sheet.max_row + 1):
    state = sheet[f"B{row}"].value
    county = sheet[f"C{row}"].value
    population = sheet[f"D{row}"].value
    info.setdefault(state, {})
    info[state].setdefault(county, {"tracts": 0, "population": 0})
    info[state][county]["tracts"] += 1
    info[state][county]["population"] += population

with open("extra_files/census.json", "w") as f:
    json.dump(info, f)

print("Population per state")
# Use of list comprehension
population_x_state = {
    state: sum([county_info["population"] for county_info in stats.values()]) for state, stats in info.items()
}
pprint.pprint(population_x_state)

# Create a new Excel file this info. Also add a Pie chart
title_font = Font(size=14, bold=True, color="FFFF0000")
state_wb = openpyxl.Workbook()
state_sheet = state_wb.active
state_sheet.title = "Information per state"
state_sheet["A1"] = "State"
state_sheet["A1"].font = title_font
state_sheet["B1"] = "Population"
state_sheet["B1"].font = title_font
state_sheet.freeze_panes = "A2"
states = population_x_state.keys()
sorted(states)
i = 0
for i, state in enumerate(states):
    state_sheet[f"A{i + 2}"] = state
    state_sheet[f"B{i + 2}"] = population_x_state[state]
state_sheet[f"B{len(states) + 2}"] = f"=SUM(B2:B{len(states) + 1})"

pie = openpyxl.chart.PieChart()
labels = openpyxl.chart.Reference(state_sheet, min_col=1, min_row=2, max_row=len(states) + 1)
data = openpyxl.chart.Reference(state_sheet, min_col=2, min_row=2, max_row=len(states) + 1)
pie.add_data(data, titles_from_data=False)
pie.set_categories(labels)

pie.title = "States"
state_sheet.add_chart(pie, "D5")
state_wb.save("extra_files/info_states.xlsx")
