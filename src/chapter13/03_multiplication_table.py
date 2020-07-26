#! /usr/bin/env python3
# Multiplication Table Maker
import argparse
import openpyxl
from openpyxl.styles import Font

print("Multiplication Table Maker")
parser = argparse.ArgumentParser()
parser.add_argument("number", help="create an number×number multiplication table in an Excel spreadsheet", type=int)
args = parser.parse_args()
print(args.number ** 2)

wb = openpyxl.Workbook()
sheet = wb.active

title_font = Font(bold=True)
for y in range(1, args.number + 1):
    sheet.cell(row=1, column=y + 1).value = y
    sheet.cell(row=1, column=y + 1).font = title_font

for x in range(1, args.number + 1):
    sheet.cell(row=x + 1, column=1).value = x
    sheet.cell(row=x + 1, column=1).font = title_font

for y in range(1, args.number + 1):
    for x in range(1, args.number + 1):
        sheet.cell(row=x + 1, column=y + 1).value = x * y

wb.save("extra_files/multiplication_table.xlsx")
