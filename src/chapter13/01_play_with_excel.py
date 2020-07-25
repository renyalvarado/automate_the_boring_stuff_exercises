#! /usr/bin/env python3
# Playing with Excel Library
import pprint
import openpyxl

print("Playing with Excel Library")
# I copied the Excel files that I previously downloaded
# from https://www.nostarch.com/download/Automate_the_Boring_Stuff_onlinematerials.zip
# into this directory
wb = openpyxl.load_workbook("extra_files/example.xlsx")
sheet = wb["Sheet1"]
cell = sheet["A1"]
print(f"{cell.coordinate}: {cell.value}")
pprint.pprint([[str(sheet.cell(row=i, column=j).value)
                for j in range(1, sheet.max_column + 1)] for i in range(1, sheet.max_row + 1)])
